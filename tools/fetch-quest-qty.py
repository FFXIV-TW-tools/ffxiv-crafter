#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""fetch-quest-qty.py — 抓「生活職業任務所需材料」試算表，產 tools/job-quest-qty.json。

輸出兩塊：`jobs`＝各職各等級的**交付數量**；`vendors`＝素材的**商人地點與單價**
（試算表「採集材料」欄，形如 `[LV05木](北黑-私語東)#商人名` 換行 `楓木原木/2G`）。

**為什麼要有這一支**：職業任務的「要交幾個」在台服解包裡找不到權威欄位
（Quest.CountableNum 全是 255 哨兵值），而遊戲裡確實有多件的任務（木工 Lv10 梣木木材 ×12）。
Owner 提供的社群試算表有這份數量，故以它為**數量的來源**——但**只當數量的來源**：
任務、交付物名稱、職業對照一律仍以解包為準（DRY 鐵則）。

**對帳規則（build-data.py 那側執行）**：只有「試算表的物品名與解包的交付物名**完全相同**」才採用數量。
名稱對不上的一律留空、UI 顯示「數量未知」——不做異體字/別名的模糊比對：
試算表用的是社群慣用名（「羅敏薩鳀魚」vs 台服官方「羅敏薩鯷魚」、「公主鱒魚」vs「公主鱒」），
猜對了沒人看得出來，猜錯了會讓採購量整批偏掉而畫面完全正常＝零回饋訊號。

用法：py -3.11 tools/fetch-quest-qty.py（需 openpyxl；偶爾手動跑，不進每次 build）
"""
import io, json, os, re, sys, urllib.request

for _s in (sys.stdout, sys.stderr):
    try: _s.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError): pass

SHEET_ID = "1St6NDm3nLERWeFJgqBS_uLZqxA91bjuMP3VwtLTGbFQ"
URL = "https://docs.google.com/spreadsheets/d/%s/export?format=xlsx" % SHEET_ID
HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "job-quest-qty.json")
# 「<職業>NN級」→ 等級；試算表每個職業一張工作表，列＝一個任務
LV_RE = re.compile(r"(\d+)\s*級")
# 「[LV05木](北黑-私語東/中黑-翡翠北)#海城木材商 ⏎ 楓木原木/2G(魚餌)」
# → 等級/類別、地點、商人（可能沒有）、物品名、單價（沒有＝只能採、沒人賣）
VENDOR_RE = re.compile(
    r"^\[LV(\d+)([^\]]*)\]\(([^)]*)\)(?:#(\S+))?\s*\n\s*([^/\n(]+?)(?:/(?:#(\S+?))?(\d+)G)?(?:\(|$|\n)",
    re.S)
# 「楓木木材×1」「橡木長弓୭×1」：୭ 是試算表自己的 HQ 記號，不屬於物品名
ITEM_RE = re.compile(r"^\s*([^（(×]+?)[୭\s]*×\s*(\d+)")


def main():
    try:
        import openpyxl
    except ImportError:
        print("✗ 需要 openpyxl：py -3.11 -m pip install openpyxl", file=sys.stderr); sys.exit(1)
    print("↓ 下載試算表…")
    # 帶 UA：不帶的話 Google 偶爾回 400（實測 2026-08-09），而那是「下載失敗」不是「表變了」，
    # 錯誤訊息卻長得像資料問題 → 容易誤判。
    req = urllib.request.Request(URL, headers={"User-Agent": "Mozilla/5.0 (ffxiv-crafter build tool)"})
    raw = urllib.request.urlopen(req, timeout=60).read()
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    # 首頁那張「地名→縮寫」對照（試算表為了排版把地名縮成兩字：黑衣森林北部林區→北黑）。
    # 不還原的話商人地點會是「北黑-私語北」，看得懂的人只有原作者。成對欄位：(全名, 縮寫)。
    areas = {}
    if wb.sheetnames:
        for row in wb[wb.sheetnames[0]].iter_rows(min_row=1, max_row=15, values_only=True):
            for i in range(0, len(row) - 1, 2):
                full, abbr = row[i], row[i + 1]
                if isinstance(full, str) and isinstance(abbr, str) and full.strip() and abbr.strip():
                    areas[abbr.strip()] = full.strip()
    data, rows_seen, vendors, cells, parsed = {}, 0, {}, 0, 0
    for name in wb.sheetnames:
        ws = wb[name]
        per_lv = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            quest, items = row[0], row[1]
            if not quest or not items:
                continue
            m = LV_RE.search(str(quest))
            if not m:
                continue
            got = {}
            for line in str(items).split("\n"):
                mm = ITEM_RE.match(line)
                if mm:
                    got[mm.group(1).strip()] = int(mm.group(2))
            if got:
                per_lv.setdefault(m.group(1), {}).update(got)
                rows_seen += 1
        # 商人/採集地點：獨立於任務列（同一張表的另一欄），故另掃一遍
        hdr = [c.value for c in ws[1]]
        if "採集材料" in (hdr or []):
            col = hdr.index("採集材料")
            for row in ws.iter_rows(min_row=2, values_only=True):
                cell = row[col] if col < len(row) else None
                if not cell:
                    continue
                cells += 1
                mm = VENDOR_RE.match(str(cell).strip())
                if not mm:
                    continue                       # 自由格式的備註列（精選素材說明等）→ 跳過，不硬解
                parsed += 1
                lv, kind, loc, npc1, item, npc2, price = mm.groups()
                if not price:
                    continue                       # 沒價格＝只能採、沒有商人賣 → 不進商人表
                vendors[item.strip()] = {"loc": (loc or "").strip(), "npc": (npc1 or npc2 or "").strip(),
                                         "price": int(price)}
        if per_lv:
            data[name] = per_lv
    if len(data) < 11:                                  # 11 職都要在；少了代表試算表結構改了 → 別靜默產半份
        print("✗ 只解析到 %d 個職業工作表（預期 ≥11）：%s" % (len(data), list(data)), file=sys.stderr)
        sys.exit(1)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump({"source": "https://docs.google.com/spreadsheets/d/%s/edit" % SHEET_ID,
                   "note": "社群整理（原始參考：灰機 Wiki）。只供『交付數量』與『商人地點/單價』；"
                           "任務、物品名、以及「這件東西到底有沒有 NPC 賣」一律以台服解包為準。",
                   "jobs": data, "vendors": vendors, "areas": areas}, f, ensure_ascii=False, indent=1)
    print("✓ job-quest-qty.json：%d 個職業 / %d 列任務 / %d 筆商人資訊 / %d 條地名縮寫"
          "（採集材料欄 %d 格，解析 %d）" % (len(data), rows_seen, len(vendors), len(areas), cells, parsed))


if __name__ == "__main__":
    main()
